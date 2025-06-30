import xml.etree.ElementTree as ET
import pandas as pd
import re
from datetime import datetime
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
import tempfile

def normalize_name(name):
    if name:
        name = re.sub(r"KB\d{7}[:\-]?\s*", "", name)
        name = re.sub(r"\s*\((January|February|March|April|May|June|July|August|September|October|November|December)\s+\d{4}\)", "", name)
        return name.strip()
    return name

def get_all_possible_tags(root):
    tags = set()
    for item in root.findall(".//ReportItem"):
        for elem in item:
            tags.add(elem.tag)
    return sorted(tags)

def get_hostname(report_host):
    for tag in report_host.findall(".//tag"):
        if tag.attrib.get("name") == "host-fqdn":
            return tag.text
    return ""

def parse_nessus(nessus_file, selected_columns=None, mode="standard"):
    tree = ET.parse(nessus_file)
    root = tree.getroot()

    results = []
    seen = set()
    report_date = datetime.today().strftime("%Y-%m-%d")

    for report_host in root.findall(".//ReportHost"):
        ip = report_host.get("name", "unknown")
        hostname = get_hostname(report_host)

        for item in report_host.findall(".//ReportItem"):
            risk_raw = item.findtext("risk_factor", default="").strip().lower()
            if risk_raw in ("none", ""):
                continue
            risk = risk_raw.capitalize()

            port = item.attrib.get("port", "")
            protocol = item.attrib.get("protocol", "").lower()
            svc_name = item.attrib.get("svc_name", "")
            service = f"{protocol}/{svc_name}" if svc_name else protocol

            name = normalize_name(item.attrib.get("pluginName", "") or item.findtext("plugin_name", ""))
            description = item.findtext("description", "")
            solution = item.findtext("solution", "")

            key = (ip.strip(), port.strip(), name.strip(), risk)
            if key in seen:
                continue
            seen.add(key)

            entry = {
                "Risk": risk,
                "Host": ip,
                "Hostname": hostname,
                "Protocol/Service": service,
                "Port": port,
                "Name": name,
                "Description": description,
                "Solution": solution,
                "Date": report_date
            }

            if mode != "standard" and selected_columns:
                for col in selected_columns:
                    entry[col] = item.findtext(col, "")

            results.append(entry)

    df = pd.DataFrame(results)
    df.sort_values(by=["Risk", "Host", "Port"], inplace=True)
    return df

if __name__ == "__main__":
    import sys

    print("\U0001F527 Nessus Rapor İşleme Aracı")
    print("1. Standart çıktı (Risk, IP, Hostname, Port, Name, Description, Solution)")
    print("2. Standart + Seçilebilir Ekstra Alanlar")
    print("3. Tüm Sütunları Kendim Belirlemek İstiyorum")

    choice = input("\nSeçiminiz (1 / 2 / 3): ").strip()
    nessus_file = input("\U0001F4C4 .nessus dosya yolu: ").strip()
    output_file = input("\U0001F4C1 Excel çıktı adı (varsayılan: output.xlsx): ").strip() or "output.xlsx"

    if choice == "1":
        df = parse_nessus(nessus_file, mode="standard")
    elif choice == "2":
        tree = ET.parse(nessus_file)
        root = tree.getroot()
        suggested_columns = [
            "pluginID", "pluginFamily", "plugin_output", "synopsis",
            "script_version", "cve", "cvss3_base_score", "exploit_available",
            "exploited_by_malware", "patch_publication_date"
        ]
        print("\n\U0001F4CC Seçilebilir Ekstra Alanlar:")
        print(" ".join(suggested_columns))
        selected = input("➕ Eklemek istediklerinizi boşlukla ayırarak yazın: ").strip().split()
        df = parse_nessus(nessus_file, selected, mode="extended")
    elif choice == "3":
        tree = ET.parse(nessus_file)
        root = tree.getroot()
        all_tags = get_all_possible_tags(root)
        print("\n\U0001F4DA Tüm Mevcut Tag'ler:")
        print(" ".join(all_tags))
        selected = input("➕ Eklemek istediklerinizi boşlukla ayırarak yazın: ").strip().split()
        df = parse_nessus(nessus_file, selected, mode="custom")
    else:
        print("❌ Geçersiz seçim.")
        sys.exit(1)

    df.to_excel(output_file, index=False, engine="openpyxl")
    print(f"\n✅ Excel dosyası oluşturuldu: {output_file}")

    # Grafiksel İstatistikler
    risk_counts = df["Risk"].value_counts().reindex(["Critical", "High", "Medium", "Low"]).fillna(0).astype(int)
    top_vulns = df["Name"].value_counts().head(5)

    # Geçici dosyalara grafik kaydet
    with tempfile.NamedTemporaryFile(suffix="_risk.png", delete=False) as tmp_risk:
        plt.figure(figsize=(6, 4))
        risk_counts.plot(kind='bar', color='tomato')
        plt.title("Zafiyetlerin Risk Seviyesi Dağılımı")
        plt.xlabel("Risk Seviyesi")
        plt.ylabel("Adet")
        plt.xticks(rotation=0)
        plt.tight_layout()
        plt.savefig(tmp_risk.name)
        risk_img_path = tmp_risk.name

    with tempfile.NamedTemporaryFile(suffix="_topvulns.png", delete=False) as tmp_top:
        plt.figure(figsize=(8, 4))
        top_vulns.plot(kind='barh', color='skyblue')
        plt.title("En Sık Görülen Zafiyetler (İlk 5)")
        plt.xlabel("Adet")
        plt.ylabel("Zafiyet İsmi")
        plt.tight_layout()
        plt.savefig(tmp_top.name)
        topvulns_img_path = tmp_top.name

    # Excel'e grafik ekle
    wb = load_workbook(output_file)
    ws = wb.create_sheet(title="Grafikler")
    ws['A1'] = "Zafiyetlerin Risk Seviyesi Dağılımı"
    img1 = XLImage(risk_img_path)
    img1.anchor = 'A2'
    ws.add_image(img1)

    ws['A20'] = "En Sık Görülen Zafiyetler (İlk 5)"
    img2 = XLImage(topvulns_img_path)
    img2.anchor = 'A21'
    ws.add_image(img2)

    wb.save(output_file)
    print("✅ Grafikler Excel dosyasına 'Grafikler' sayfası olarak eklendi.")
